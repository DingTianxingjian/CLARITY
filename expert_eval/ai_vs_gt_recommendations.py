"""
AI (Claude) vs GT treatment recommendations — generates comparison Excel.
Run with the mewm env python.
"""
import sys
sys.path.insert(0, 'CLARITY/expert_eval')
from generate_latex import load_clinical, find_pairs, select_diverse, fmt_actions
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from pathlib import Path

# ─── AI recommendations (Claude, ignoring GT) ────────────────────────────────
# Each entry:  (imaging_summary, clinical_reasoning, ai_recommendation)
AI_RECS = {
    # pid: (mri_obs, reasoning, recommendation)
    "PatientID_0210": (
        "Small ring-enhancing lesion right parietal/thalamic on T1c with central necrosis; "
        "extensive T2/FLAIR edema. Enhancement pattern inconsistent with grade 2 — behaves as GBM.",
        "IDH-WT 'grade 2' astrocytoma is biologically GBM-equivalent. Patient already progressed "
        "through RT 60Gy + TMZ + Avastin/Everolimus (very aggressive 1st line). "
        "Salvage options limited; re-challenge with Avastin+TMZ or switch to lomustine.",
        "Bevacizumab 10 mg/kg q14d + Temozolomide 150 mg/m² d1-5 q28d x6 cycles. "
        "If confirmed Avastin-refractory: Lomustine (CCNU) 90 mg/m² q6w. Clinical trial strongly preferred."
    ),
    "PatientID_0207": (
        "Large irregular ring-enhancing mass right frontal/temporal with necrotic core on T1c; "
        "extensive T2 edema crossing midline. Aggressive imaging despite IDH-mutant status.",
        "IDH-mutant grade 4 astrocytoma (WHO 2021). TP53-mutant, ATRX-WT, no 1p/19q codeletion. "
        "Newly diagnosed, no prior treatment. Stupp protocol is standard. IDH-mutant confers "
        "modestly better prognosis vs IDH-WT GBM; MGMT unknown but TMZ still recommended.",
        "RT 60 Gy/30 fr + concurrent TMZ 75 mg/m²/day → adjuvant TMZ x6 cycles "
        "(150-200 mg/m² d1-5 q28d) + Optune TTF after CRT phase."
    ),
    "PatientID_0209": (
        "Irregular ring-enhancing mass left frontal/parietal with central necrosis on T1c; "
        "satellite lesion possible; significant surrounding edema. Post-surgical cavity visible.",
        "IDH-mutant grade 4 + MGMT methylated — most favorable GBM subgroup. TP53-mutant, "
        "ATRX-loss, no 1p/19q codeletion. Newly diagnosed. MGMT methylation predicts strong "
        "TMZ benefit; consider extended adjuvant TMZ beyond 6 cycles.",
        "RT 60 Gy/30 fr + concurrent TMZ 75 mg/m²/day → adjuvant TMZ x12 cycles "
        "(150-200 mg/m² d1-5 q28d; MGMT methylation supports extension) + Optune TTF."
    ),
    "PatientID_0188": (
        "Multifocal or post-surgical enhancing lesion posterior parietal; bilateral T2/FLAIR "
        "signal suggesting infiltrative disease. Complex imaging pattern.",
        "IDH-WT grade 3 astrocytoma + MGMT methylated. IDH-WT grade 3 without +7/-10 is rare; "
        "behaves aggressively. CATNON data supports RT + TMZ for non-codeleted grade 3. "
        "MGMT methylation predicts TMZ response. 63yo — full Stupp dosing appropriate if PS good.",
        "RT 59.4 Gy/33 fr + concurrent TMZ 75 mg/m²/day → adjuvant TMZ x12 cycles "
        "(CATNON regimen for IDH-WT/uncertain non-codeleted grade 3)."
    ),
    "PatientID_0242": (
        "No enhancement on T1c; midline/callosal T2-bright non-enhancing lesion with well-defined "
        "borders on T2/FLAIR. Classic low-grade oligodendroglioma appearance.",
        "Oligodendroglioma grade 2, 1p/19q codeleted, MGMT methylated. Pre-scan treatment: "
        "RT 54 Gy + PCV x6 q42d already ongoing — mid-protocol. No progression (first progression "
        "day 1926). Simply continue the current protocol.",
        "Continue adjuvant PCV x6 cycles q42d per RTOG 9802 protocol. No modification needed."
    ),
    "PatientID_0036": (
        "Large heterogeneously enhancing right frontal/temporal/parietal mass with irregular "
        "borders and partial necrosis on T1c; extensive T2 edema. Post-surgical changes visible.",
        "GBM IDH-WT, MGMT unknown. Just completed concurrent CRT phase. Entering standard "
        "adjuvant phase. MGMT testing should guide decision; unknown status → proceed with "
        "standard adjuvant TMZ + Optune (EF-14 data).",
        "Adjuvant TMZ x6 cycles (150-200 mg/m² d1-5 q28d) + Optune TTF continuous. "
        "Obtain MGMT status to guide future therapy decisions."
    ),
    "PatientID_0236": (
        "Minimal T1c enhancement; small central/deep midline lesion (thalamic/basal ganglia?) "
        "bright on T2w. Deep eloquent location. Small volume, non-enhancing.",
        "Astrocytoma grade 2, IDH unknown, 25yo male. Progression at TP1 (day 105). "
        "Deep eloquent location limits surgical options. IDH likely mutant at this age. "
        "RT alone (GT choice) may be insufficient — grade 2 with progression warrants chemo. "
        "RT dose per GT (50.4 Gy) appropriate for low-grade/eloquent location.",
        "RT 50.4 Gy/28 fr + concurrent/adjuvant TMZ x6 cycles q28d. "
        "Obtain IDH status: if IDH-mutant + 1p/19q codeleted → switch to PCV x6 q42d adjuvant."
    ),
    "PatientID_0195": (
        "Small, minimally enhancing right temporal/inferior lesion on T1c; small T2/FLAIR "
        "signal with minimal edema. Classic low-grade appearance, small volume.",
        "IDH-mutant grade 2 astrocytoma, MGMT unmethylated, 1p/19q intact. Age 38 "
        "(borderline high-risk). No prior treatment, no progression. For high-risk LGG "
        "(age ≥40 or large tumor), RT + chemotherapy is indicated. At 38, treat actively. "
        "MGMT unmethylated limits TMZ benefit but standard of care includes it. "
        "GT used 60 Gy (GBM dosing) — 54 Gy is more appropriate for grade 2.",
        "RT 54 Gy/30 fr + concurrent TMZ → adjuvant TMZ x6 cycles q28d. "
        "Consider RTOG 9802 PCV if progression occurs on TMZ (IDH-mutant, 1p/19q intact)."
    ),
    "PatientID_0275": (
        "Tiny, nearly imperceptible T1c enhancement; very small right frontal cortical "
        "lesion on T2w/FLAIR. Classic low-grade oligodendroglioma — small volume, minimal enhancement.",
        "Oligodendroglioma grade 2, IDH-mutant, 1p/19q codeleted, MGMT methylated, TERT-mutant. "
        "Pre-scan: RT 54 Gy alone. GT shows RT 54 Gy again — likely data artifact; "
        "more plausible interpretation is RT completed, now adding adjuvant PCV. "
        "RTOG 9802 shows RT + PCV superior to RT alone for this histology.",
        "Adjuvant PCV x6 cycles q42d following RT completion. "
        "Do NOT repeat RT without confirmed progression. Standard RTOG 9802 protocol."
    ),
    "PatientID_0072": (
        "Large ring-enhancing lesion left frontal/parietal with central necrosis and "
        "necrotic core on T1c. Classic GBM ring enhancement. Significant T2 edema with "
        "mass effect and midline shift. Large tumor burden.",
        "GBM IDH-WT, MGMT unmethylated. Just completed concurrent CRT. MGMT unmethylated "
        "predicts reduced TMZ benefit, but adjuvant TMZ remains standard of care. "
        "Optune adds PFS benefit per EF-14 regardless of MGMT status. "
        "Bevacizumab not recommended 1st-line (no OS benefit in AVAglio/RTOG 0825).",
        "Adjuvant TMZ x6 cycles (150-200 mg/m² d1-5 q28d) + Optune TTF continuous. "
        "At first progression (MGMT unmet): bevacizumab or lomustine-based salvage."
    ),
    "PatientID_0235": (
        "Large heterogeneous enhancing left hemisphere mass with necrosis on T1c; "
        "extensive T2 edema and infiltration. Imaging is incompatible with grade 2 "
        "low-grade glioma — aggressive GBM-like appearance.",
        "IDH-WT 'grade 2' with rapid progression (day 57, type 3 = malignant transformation). "
        "Imaging shows aggressive enhancement with necrosis. IDH-WT + this MRI = effectively GBM. "
        "GT choice of PCV alone (no RT) is inadequate — RT is essential for IDH-WT aggressive glioma. "
        "PCV without RT may be used for IDH-mutant oligodendroglioma, not IDH-WT with GBM imaging.",
        "RT 60 Gy/30 fr + concurrent TMZ 75 mg/m²/day → adjuvant TMZ x6 cycles. "
        "Treat as de facto GBM given IDH-WT status + aggressive imaging + early malignant progression. "
        "PCV alone without RT is insufficient."
    ),
    "PatientID_0033": (
        "Moderately sized left frontal/parietal enhancing lesion on T1c with heterogeneous "
        "enhancement. T2/FLAIR shows edema. Moderate tumor burden post-CRT changes.",
        "GBM IDH-WT, MGMT methylated (favorable). Just completed concurrent CRT (dose unknown). "
        "MGMT methylation strongly predicts TMZ benefit. Standard adjuvant TMZ + Optune. "
        "Consider extending TMZ beyond 6 cycles given MGMT methylation (some centers use 12 cycles).",
        "Adjuvant TMZ x6 cycles (150-200 mg/m² d1-5 q28d) + Optune TTF continuous. "
        "Given MGMT methylation, consider extension to x12 cycles if tolerated and no progression."
    ),
    "PatientID_0272": (
        "Extremely large bilateral/extensive enhancing mass on T1c occupying majority of brain "
        "parenchyma. Post-surgical changes. Massive tumor burden. Most extensive imaging of cohort.",
        "GBM grade 4, IDH unknown, MGMT unmethylated, CDKN2A/B deletion, ATRX loss. "
        "Completed concurrent CRT. Despite large burden, no progression in interval (first prog: unknown). "
        "MGMT unmethylated limits TMZ benefit. Optune appropriate (EF-14 data). "
        "Aggressive surveillance given volume.",
        "Adjuvant TMZ x6 cycles (q28d) + Optune TTF continuous. "
        "Given MGMT unmethylated and massive burden, plan early escalation to bevacizumab at first "
        "radiographic progression. Early MDT discussion re: goals of care."
    ),
    "PatientID_0085": (
        "Small ring-enhancing lesion right thalamic/deep central region on T1c. "
        "Deep eloquent location. Small enhancement with necrotic center. Perilesional edema.",
        "GBM IDH-WT, MGMT unmethylated, young (37yo). Hypofractionated RT 40 Gy/15 fr chosen "
        "(likely due to deep eloquent location), already 2/6 adjuvant TMZ cycles completed. "
        "MGMT unmethylated reduces TMZ efficacy. Optune feasible. "
        "At progression: bevacizumab or lomustine (CCNU). Clinical trial preferred given young age.",
        "Continue adjuvant TMZ x4 more cycles (total 6) + Optune TTF. "
        "At progression: Bevacizumab 10 mg/kg q14d or CCNU 90 mg/m² q6w. "
        "Strongly recommend clinical trial enrollment given age 37."
    ),
    "PatientID_0204": (
        "Large enhancing right hemisphere mass with mass effect and midline shift on T1c; "
        "multiple enhancement foci with central necrotic areas. Extensive T2/FLAIR infiltration. "
        "Large tumor despite young age and IDH-mutant status.",
        "IDH-mutant grade 4 astrocytoma, MGMT methylated, TP53-mutant, ATRX-loss. "
        "Very young (24yo). Best-prognosis molecular subgroup (IDH-mut + MGMT-met). "
        "Standard Stupp protocol mandatory. MGMT methylation supports extended TMZ. "
        "Consider IDH inhibitor (ivosidenib) in later lines given IDH1 mutation.",
        "RT 60 Gy/30 fr + concurrent TMZ 75 mg/m²/day → adjuvant TMZ x12 cycles "
        "(MGMT methylated; extended per institutional protocol) + Optune TTF. "
        "Clinical trial with IDH-targeted agent for recurrence."
    ),
    "PatientID_0250": (
        "Very large heterogeneously enhancing left hemisphere mass with massive surrounding "
        "edema on T1c; almost entire left hemisphere affected with significant midline shift. "
        "Extreme tumor burden. Left-sided likely with functional deficits.",
        "GBM IDH-WT, MGMT methylated. Completed concurrent CRT (60 Gy + TMZ). Early progression "
        "(108 days from diagnosis — 4 days after TP1). Rapid progression despite standard CRT "
        "and MGMT methylation. GT chooses brachytherapy alone. "
        "For recurrent GBM: bevacizumab ± lomustine preferred. Given MGMT-met, lomustine re-challenge "
        "has rationale. Brachytherapy is one option but has limited evidence.",
        "Bevacizumab 10 mg/kg q14d + Lomustine (CCNU) 90 mg/m² q6w (BELOB regimen; "
        "MGMT methylation supports alkylating re-challenge). "
        "Goals-of-care discussion given rapid progression and large burden. "
        "Palliative care integration. SRS boost if PS allows and focal recurrence confirmed."
    ),
    "PatientID_0054": (
        "Large right frontal/parietal heterogeneously enhancing mass with areas of necrosis; "
        "significant T2 edema. Substantial tumor burden with irregular borders. "
        "Post-CRT changes coexist.",
        "GBM IDH-WT, MGMT methylated, 65yo female. Completed RT 60 Gy + TMZ + Optune. "
        "Very rapid progression (99 days — day after TP1). Progressive despite Optune + CRT. "
        "MGMT methylated but still progressed rapidly. GT continues TMZ x2 + Optune — modest. "
        "For post-CRT rapidly progressive GBM: bevacizumab improves PFS and QoL. "
        "Adding lomustine given MGMT methylation.",
        "Bevacizumab 10 mg/kg q14d + Lomustine 90 mg/m² q6w (MGMT-met favors alkylating rechallenge). "
        "Optune can continue if PS allows. "
        "Early palliative care integration; goals-of-care discussion given age 65 + rapid progression."
    ),
    "PatientID_0267": (
        "Left temporal/basal ganglia heterogeneous enhancement adjacent to eloquent structures; "
        "moderate T2 edema. Medium-sized lesion near functional territory. "
        "Post-surgical changes.",
        "GBM IDH-WT, MGMT methylated, EGFR amplified, TERT mutant, 66yo male. "
        "Completed concurrent CRT. No progression yet (first prog day 264). "
        "GT skips adjuvant TMZ entirely → uses Bevacizumab + Optune as adjuvant. "
        "This is NON-STANDARD: MGMT methylation strongly mandates adjuvant TMZ continuation. "
        "Bevacizumab does not improve OS in first-line adjuvant (AVAglio). "
        "EGFR amp: no approved targeted agent currently.",
        "Adjuvant TMZ x6 cycles (150-200 mg/m² d1-5 q28d) + Optune TTF continuous. "
        "MGMT methylated = strong indication for TMZ. Bevacizumab reserved for confirmed progression. "
        "No evidence for EGFR-directed therapy in GBM outside trial."
    ),
    "PatientID_0038": (
        "Left frontal heterogeneous enhancement with post-surgical cavity on T1c; "
        "significant T2/FLAIR signal indicating residual/infiltrative disease. "
        "Moderate tumor burden.",
        "GBM IDH-WT, MGMT unmethylated, Gliadel wafer placed at resection. "
        "Completed concurrent CRT (RT 60 Gy + TMZ + local carmustine via wafer). "
        "Entering adjuvant phase. MGMT unmethylated limits systemic TMZ benefit, "
        "but Gliadel provided local BCNU. Standard adjuvant TMZ + Optune is SoC.",
        "Adjuvant TMZ x6 cycles (150-200 mg/m² d1-5 q28d) + Optune TTF continuous. "
        "Gliadel already provided local chemotherapy; systemic TMZ continuation is standard. "
        "At progression (MGMT unmet): bevacizumab or clinical trial."
    ),
    "PatientID_0087": (
        "MASSIVE enhancing mass occupying inferior brain on T1c — most extensive imaging in cohort. "
        "Near-total brain involvement with central necrosis. Extensive edema. "
        "Critically large tumor burden suggesting very poor functional status.",
        "GBM IDH-WT, MGMT unmethylated, 61yo female. Completed CRT (dose unknown). "
        "Progression at day 133 (4 days after TP1). Essentially no response to CRT. "
        "MGMT unmethylated — limited benefit from further TMZ. Massive tumor. "
        "GT repeats RT + TMZ (same failed regimen) — inadequate rationale. "
        "Bevacizumab is preferred for edema control and PFS in this setting. "
        "Re-RT may help focally, but tumor extent makes focal re-RT impossible here. "
        "Goals-of-care discussion is urgent given this presentation.",
        "Bevacizumab 10 mg/kg q14d (edema control + PFS benefit). "
        "Supportive corticosteroids (dexamethasone) for symptom management. "
        "Mandatory goals-of-care discussion given massive tumor extent, rapid progression, "
        "MGMT unmethylated status, and poor prognosis. "
        "Palliative care integration. Re-RT not feasible given tumor extent."
    ),
}

# ─── Build Excel ─────────────────────────────────────────────────────────────

def thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr_style(cell, bg="1E3C78", fg="FFFFFF", size=10, bold=True):
    cell.font = Font(name="Calibri", bold=bold, size=size, color=fg)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center",
                               wrap_text=True)
    cell.border = thin_border()

def data_style(cell, bg=None, bold=False, size=9, wrap=True, halign="left"):
    cell.font = Font(name="Calibri", bold=bold, size=size)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=halign, vertical="top", wrap_text=wrap)
    cell.border = thin_border()


def build_excel(cases_meta, out_path):
    wb = openpyxl.Workbook()

    # ── Sheet 1: full comparison ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "AI vs GT"

    headers = [
        "Case", "Patient ID", "Age/Sex", "Diagnosis", "WHO Grade",
        "Key Genomics",
        "Prior Treatment (pre-scan)",
        "MRI Impression",
        "Clinical Reasoning (AI)",
        "AI Recommendation",
        "GT Treatment",
        "Match Level",
        "Notes"
    ]
    ws.append(headers)
    for ci, h in enumerate(headers, 1):
        hdr_style(ws.cell(1, ci))

    MATCH_COLORS = {
        "Full match":       "C6EFCE",  # green
        "Partial match":    "FFEB9C",  # yellow
        "Disagree":         "FFC7CE",  # red
        "N/A (salvage)":    "DDEBF7",  # blue
    }

    for i, meta in enumerate(cases_meta):
        pid = meta["pid"]
        rec = AI_RECS.get(pid, ("—", "—", "—"))
        mri_obs, reasoning, ai_rec = rec

        gt_lines = meta["gt_lines"]
        gt_text = "\n".join(
            l.replace("\\textbf{", "").replace("}", "")
             .replace("\\textit{", "").replace("$\\times$", "x")
             .replace("\\,", " ")
            for l in gt_lines
        )

        # rough match level
        ai_low = ai_rec.lower()
        gt_low = gt_text.lower()
        if "salvage" in reasoning.lower() or "palliative" in reasoning.lower():
            match = "N/A (salvage)"
        elif any(k in ai_low and k in gt_low for k in
                 ["tmz", "temozolomide", "pcv", "bevacizumab", "avastin", "optune"]):
            core_match = sum(
                1 for k in ["tmz","pcv","bevacizumab","avastin","optune","lomustine","rt"]
                if k in ai_low and k in gt_low
            )
            match = "Full match" if core_match >= 2 else "Partial match"
        else:
            match = "Disagree"

        row = [
            meta["case_num"],
            pid,
            f"{meta['age']} / {meta['sex']}",
            meta["diagnosis"],
            meta["who_grade"],
            meta["genomics_str"].replace("\\textit{", "").replace("}", ""),
            "\n".join(
                l.replace("\\textbf{", "").replace("}", "")
                 .replace("$\\times$", "x").replace("\\,", " ")
                 .replace("\\textit{", "")
                for l in meta["pre_lines"]
            ),
            mri_obs,
            reasoning,
            ai_rec,
            gt_text,
            match,
            "",
        ]
        r = i + 2
        ws.append(row)

        bg_alt = "F5F8FF" if i % 2 == 0 else "FFFFFF"
        for ci, val in enumerate(row, 1):
            cell = ws.cell(r, ci)
            if ci == 12:  # Match Level
                data_style(cell, bg=MATCH_COLORS.get(match, bg_alt),
                           bold=True, halign="center")
            elif ci in (1, 2, 3, 4, 5):
                data_style(cell, bg=bg_alt, bold=(ci==1), halign="center")
            else:
                data_style(cell, bg=bg_alt)

    # column widths
    widths = [6, 18, 10, 20, 8, 32, 28, 38, 42, 42, 28, 14, 18]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 30
    for r in range(2, len(cases_meta)+2):
        ws.row_dimensions[r].height = 90

    ws.freeze_panes = "A2"

    # ── Sheet 2: summary stats ────────────────────────────────────────────────
    ss = wb.create_sheet("Match Summary")
    ss.append(["Match Category", "Count", "% of 20 cases", "Interpretation"])
    hdr_style(ss.cell(1,1)); hdr_style(ss.cell(1,2))
    hdr_style(ss.cell(1,3)); hdr_style(ss.cell(1,4))

    counts = {"Full match": 0, "Partial match": 0, "Disagree": 0, "N/A (salvage)": 0}
    for i, meta in enumerate(cases_meta):
        pid = meta["pid"]
        rec = AI_RECS.get(pid, ("—","—","—"))
        ai_rec, reasoning = rec[2], rec[1]
        gt_lines = meta["gt_lines"]
        gt_text = " ".join(gt_lines).lower()
        ai_low = ai_rec.lower()
        if "salvage" in reasoning.lower() or "palliative" in reasoning.lower():
            match = "N/A (salvage)"
        elif any(k in ai_low and k in gt_text for k in
                 ["tmz","pcv","bevacizumab","avastin","optune"]):
            core = sum(1 for k in ["tmz","pcv","bevacizumab","avastin","optune","lomustine","rt"]
                       if k in ai_low and k in gt_text)
            match = "Full match" if core >= 2 else "Partial match"
        else:
            match = "Disagree"
        counts[match] += 1

    interp = {
        "Full match":    "AI and GT agree on primary agent(s) and regimen class",
        "Partial match": "AI and GT share >=1 agent but differ in dose/duration/combination",
        "Disagree":      "AI recommends fundamentally different regimen than GT",
        "N/A (salvage)": "Case involves recurrent/salvage setting where GT itself is varied",
    }
    colors = {"Full match":"C6EFCE","Partial match":"FFEB9C",
              "Disagree":"FFC7CE","N/A (salvage)":"DDEBF7"}
    for r, (cat, cnt) in enumerate(counts.items(), 2):
        ss.cell(r,1,cat); ss.cell(r,2,cnt)
        ss.cell(r,3,f"{cnt/20*100:.0f}%"); ss.cell(r,4,interp[cat])
        for ci in range(1,5):
            cell = ss.cell(r,ci)
            cell.fill = PatternFill("solid", fgColor=colors[cat])
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(horizontal="center" if ci<4 else "left",
                                       vertical="center", wrap_text=True)
            cell.border = thin_border()

    for col, w in zip("ABCD", [18,8,14,55]):
        ss.column_dimensions[col].width = w
    for r in range(1,6):
        ss.row_dimensions[r].height = 24

    wb.save(out_path)
    print(f"Saved: {out_path}")


# ─── main ─────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    patients  = load_clinical()
    pairs     = find_pairs(patients)
    selected  = select_diverse(pairs, n=20)
    cases_meta = []
    for i, c in enumerate(selected):
        from generate_latex import fmt_case_meta
        cases_meta.append(fmt_case_meta(c, i+1))

    out = Path("CLARITY/expert_eval/ai_vs_gt_comparison.xlsx")
    build_excel(cases_meta, out)
