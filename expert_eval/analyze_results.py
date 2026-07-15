"""
CLARITY vs GT Expert Evaluation — Result Analyzer
==================================================
用法：
  python analyze_results.py \
    --response1 expert1_filled.xlsx \
    --response2 expert2_filled.xlsx \
    --mapping   researcher_mapping_CONFIDENTIAL.xlsx

输出：
  - 控制台打印完整结果表
  - results_summary.xlsx
"""

import argparse, sys
import numpy as np
import openpyxl
from pathlib import Path
from sklearn.metrics import cohen_kappa_score


# ─── Load expert responses ────────────────────────────────────────────────────

def load_response(path: str) -> dict:
    """返回 {case_no: {"choice": "A"|"B"|"Equivalent", "confidence": int, "comment": str}}"""
    wb = openpyxl.load_workbook(path)
    ws = wb["Expert Response"]
    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        case_no = int(row[0])
        choice  = str(row[3]).strip() if row[3] else ""
        conf    = int(row[4]) if row[4] else None
        comment = str(row[5]).strip() if row[5] else ""
        result[case_no] = {"choice": choice, "confidence": conf, "comment": comment}
    return result


def load_mapping(path: str) -> dict:
    """返回 {case_no: {"a_is": "CLARITY"|"GT", "pid": str, ...}}"""
    wb = openpyxl.load_workbook(path)
    ws = wb["GT Mapping (CONFIDENTIAL)"]
    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        case_no = int(row[0])
        result[case_no] = {
            "pid":          row[1],
            "pre_tp":       row[2],
            "post_tp":      row[3],
            "a_is":         row[4],   # "CLARITY" or "GT"
            "b_is":         row[5],
            "gt_treatment": row[6],
            "clarity_plan": row[7],
            "survival_days": row[8],
            "event":        row[9],
        }
    return result


# ─── Decode choice ────────────────────────────────────────────────────────────

def decode_choice(choice: str, a_is: str) -> str:
    """
    Convert "A"/"B"/"Equivalent" + A_is_mapping -> "CLARITY"/"GT"/"Equivalent"
    """
    c = choice.upper().strip()
    if c in ("EQUIVALENT", "TIE", "BOTH", "C", "=", "EQUAL"):
        return "Equivalent"
    b_is = "GT" if a_is == "CLARITY" else "CLARITY"
    if c == "A":
        return a_is
    if c == "B":
        return b_is
    return "Unknown"


# ─── Metrics ──────────────────────────────────────────────────────────────────

def compute_metrics(decoded: list, label: str) -> dict:
    """
    decoded: list of "CLARITY" | "GT" | "Equivalent"
    Returns non_inferior_rate, clarity_win_rate, gt_win_rate, tie_rate
    """
    n = len(decoded)
    clarity_win = sum(1 for x in decoded if x == "CLARITY")
    gt_win      = sum(1 for x in decoded if x == "GT")
    tie         = sum(1 for x in decoded if x == "Equivalent")

    return {
        "expert":           label,
        "n_cases":          n,
        "CLARITY_win":      clarity_win,
        "GT_win":           gt_win,
        "Tie":              tie,
        "Unknown":          n - clarity_win - gt_win - tie,
        "non_inferior_%":   round((clarity_win + tie) / n * 100, 1),
        "clarity_win_%":    round(clarity_win / n * 100, 1),
        "gt_win_%":         round(gt_win / n * 100, 1),
        "tie_%":            round(tie / n * 100, 1),
    }


def cohen_kappa(decoded1: list, decoded2: list) -> float:
    labels = ["CLARITY", "GT", "Equivalent"]
    label_map = {l: i for i, l in enumerate(labels)}
    y1 = [label_map.get(x, -1) for x in decoded1]
    y2 = [label_map.get(x, -1) for x in decoded2]
    # filter Unknown
    valid = [(a, b) for a, b in zip(y1, y2) if a >= 0 and b >= 0]
    if not valid:
        return float("nan")
    a, b = zip(*valid)
    return cohen_kappa_score(list(a), list(b))


# ─── Output ──────────────────────────────────────────────────────────────────

def print_table(header: list, rows: list):
    col_w = [max(len(str(h)), max((len(str(r[i])) for r in rows), default=0))
             for i, h in enumerate(header)]
    sep = "+" + "+".join("-" * (w + 2) for w in col_w) + "+"
    fmt = "|" + "|".join(f" {{:<{w}}} " for w in col_w) + "|"
    print(sep)
    print(fmt.format(*header))
    print(sep)
    for row in rows:
        print(fmt.format(*[str(x) for x in row]))
    print(sep)


def build_summary_xlsx(case_rows: list, metrics1: dict, metrics2: dict,
                        kappa: float, out_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Case-Level Results"

    hdr = ["Case", "PID", "Pre TP", "Post TP", "A is", "B is",
           "Expert1 Choice", "Expert1 Decoded", "Expert1 Conf",
           "Expert2 Choice", "Expert2 Decoded", "Expert2 Conf",
           "Agreement?", "Survival Days", "Event"]
    ws.append(hdr)

    from openpyxl.styles import Font, PatternFill, Alignment
    fill_clarity = PatternFill("solid", fgColor="D4EDDA")  # green
    fill_gt      = PatternFill("solid", fgColor="F8D7DA")  # red
    fill_tie     = PatternFill("solid", fgColor="FFF3CD")  # yellow

    for r in case_rows:
        ws.append([r[h] for h in hdr])
        row_idx = ws.max_row
        for col_idx, key in enumerate(hdr, 1):
            cell = ws.cell(row_idx, col_idx)
            if key == "Expert1 Decoded":
                if cell.value == "CLARITY": cell.fill = fill_clarity
                elif cell.value == "GT":    cell.fill = fill_gt
                else:                       cell.fill = fill_tie
            if key == "Expert2 Decoded":
                if cell.value == "CLARITY": cell.fill = fill_clarity
                elif cell.value == "GT":    cell.fill = fill_gt
                else:                       cell.fill = fill_tie

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 16
    ws.column_dimensions["J"].width = 16
    ws.column_dimensions["K"].width = 16

    # Summary sheet
    sm = wb.create_sheet("Summary Metrics")
    sm.append(["Metric", "Expert 1", "Expert 2", "Average"])
    for key in ["non_inferior_%", "clarity_win_%", "gt_win_%", "tie_%"]:
        label = {
            "non_inferior_%": "Non-inferior to GT (%)",
            "clarity_win_%":  "CLARITY strict-win (%)",
            "gt_win_%":       "GT strict-win (%)",
            "tie_%":          "Tie / Equivalent (%)",
        }[key]
        avg = round((metrics1[key] + metrics2[key]) / 2, 1)
        sm.append([label, metrics1[key], metrics2[key], avg])

    sm.append([""])
    sm.append(["Inter-rater Cohen's kappa", round(kappa, 3), "", ""])
    sm.append(["kappa interpretation",
               "< 0.2 Slight  |  0.2-0.4 Fair  |  0.4-0.6 Moderate  |  0.6-0.8 Substantial  |  > 0.8 Almost perfect",
               "", ""])

    for col in ["A", "B", "C", "D"]:
        sm.column_dimensions[col].width = 35

    wb.save(out_path)


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--response1", required=True)
    parser.add_argument("--response2", required=True)
    parser.add_argument("--mapping",   required=True)
    parser.add_argument("--out", default="results_summary.xlsx")
    args = parser.parse_args()

    print("Loading data...")
    resp1   = load_response(args.response1)
    resp2   = load_response(args.response2)
    mapping = load_mapping(args.mapping)

    n = len(mapping)
    print(f"  {n} cases in mapping, {len(resp1)} in expert1, {len(resp2)} in expert2")

    decoded1, decoded2 = [], []
    case_rows = []

    for case_no in sorted(mapping.keys()):
        m = mapping[case_no]
        r1 = resp1.get(case_no, {})
        r2 = resp2.get(case_no, {})

        d1 = decode_choice(r1.get("choice", ""), m["a_is"])
        d2 = decode_choice(r2.get("choice", ""), m["a_is"])
        decoded1.append(d1)
        decoded2.append(d2)

        case_rows.append({
            "Case":             case_no,
            "PID":              m["pid"],
            "Pre TP":           m["pre_tp"],
            "Post TP":          m["post_tp"],
            "A is":             m["a_is"],
            "B is":             m["b_is"],
            "Expert1 Choice":   r1.get("choice", "MISSING"),
            "Expert1 Decoded":  d1,
            "Expert1 Conf":     r1.get("confidence", ""),
            "Expert2 Choice":   r2.get("choice", "MISSING"),
            "Expert2 Decoded":  d2,
            "Expert2 Conf":     r2.get("confidence", ""),
            "Agreement?":       "YES" if d1 == d2 else "NO",
            "Survival Days":    m["survival_days"],
            "Event":            m["event"],
        })

    metrics1 = compute_metrics(decoded1, "Expert 1")
    metrics2 = compute_metrics(decoded2, "Expert 2")
    kappa    = cohen_kappa(decoded1, decoded2)

    avg_non_inf = round((metrics1["non_inferior_%"] + metrics2["non_inferior_%"]) / 2, 1)
    avg_win     = round((metrics1["clarity_win_%"]  + metrics2["clarity_win_%"])  / 2, 1)

    print("\n" + "=" * 60)
    print("PRIMARY RESULT")
    print("=" * 60)
    print(f"  Non-inferior to GT:  Expert1={metrics1['non_inferior_%']}%  "
          f"Expert2={metrics2['non_inferior_%']}%  "
          f"Average={avg_non_inf}%")
    print(f"  CLARITY strict-win:  Expert1={metrics1['clarity_win_%']}%  "
          f"Expert2={metrics2['clarity_win_%']}%  "
          f"Average={avg_win}%")
    print(f"  Inter-rater kappa:   kappa={kappa:.3f}")
    print()

    hdr = ["Expert", "Non-inferior %", "CLARITY-win %", "GT-win %", "Tie %"]
    rows = [
        [metrics1["expert"], metrics1["non_inferior_%"], metrics1["clarity_win_%"],
         metrics1["gt_win_%"], metrics1["tie_%"]],
        [metrics2["expert"], metrics2["non_inferior_%"], metrics2["clarity_win_%"],
         metrics2["gt_win_%"], metrics2["tie_%"]],
        ["Average",
         avg_non_inf,
         avg_win,
         round((metrics1["gt_win_%"] + metrics2["gt_win_%"]) / 2, 1),
         round((metrics1["tie_%"]    + metrics2["tie_%"])    / 2, 1)],
    ]
    print_table(hdr, rows)

    print(f"\nCohen's kappa: {kappa:.3f}")
    if kappa >= 0.6:
        print("  -> Substantial agreement: conclusions are reliable")
    elif kappa >= 0.4:
        print("  -> Moderate agreement: consider adjudication for discordant cases")
    else:
        print("  -> Fair/poor agreement: recommend third-expert adjudication")

    out_path = Path(args.out)
    build_summary_xlsx(case_rows, metrics1, metrics2, kappa, out_path)
    print(f"\nSaved: {out_path}")

    print("\nTable for paper (Tab. 6):")
    print(f"  CLARITY (Expert 1): Non-inferior={metrics1['non_inferior_%']}%  "
          f"Strict-win={metrics1['clarity_win_%']}%")
    print(f"  CLARITY (Expert 2): Non-inferior={metrics2['non_inferior_%']}%  "
          f"Strict-win={metrics2['clarity_win_%']}%")
    print(f"  CLARITY (averaged): Non-inferior={avg_non_inf}%  Strict-win={avg_win}%")
    print(f"  Inter-rater kappa:  {kappa:.3f}")


if __name__ == "__main__":
    main()
