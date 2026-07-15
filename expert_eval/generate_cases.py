"""
CLARITY vs GT Blind Expert Evaluation - Case Generator
=======================================================
Output:
  cases/case_XX_expert.pdf     - 盲态 case 材料（给专家）
  mri_slices/                  - 每个 case 的 MRI 切片 PNG
  expert_response_sheet.xlsx   - 专家填写表（无 A/B 身份信息）
  researcher_mapping.xlsx      - 研究者保存，记录 A/B 真实身份（专家不可见）

用法：
  python generate_cases.py [--clarity_json PATH]  # 若有 CLARITY 输出 JSON，传入
  python generate_cases.py                         # 无 CLARITY 输出时，用 [CLARITY PENDING] 占位
"""

import os, sys, json, random, re, argparse
import numpy as np
import nibabel as nib
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from pathlib import Path
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fpdf import FPDF

# ─── Paths ────────────────────────────────────────────────────────────────────
ROOT         = Path(__file__).resolve().parents[2]
MRI_ROOT     = ROOT / "datasets" / "MU-Glioma-Post"
CLINICAL_JSON = ROOT / "Predictor" / "dataset" / "MU_Glioma_Post" / "clinical_latest.json"
OUT_DIR      = Path(__file__).parent
CASES_DIR    = OUT_DIR / "cases"
MRI_DIR      = OUT_DIR / "mri_slices"
CASES_DIR.mkdir(exist_ok=True)
MRI_DIR.mkdir(exist_ok=True)

NUM_CASES    = 20
RANDOM_SEED  = 2024

# ─── Genomic label map ────────────────────────────────────────────────────────
GENO_LABELS = {
    "idh1":              {0: "IDH1 WT", 1: "IDH1 mut", 2: "IDH1 unk"},
    "idh2":              {0: "IDH2 WT", 1: "IDH2 mut", 2: "IDH2 unk"},
    "mgmt_methylation":  {0: "MGMT unmethylated", 1: "MGMT methylated", 2: "MGMT unk"},
    "egfr_amp":          {0: "EGFR normal", 1: "EGFR amp", 2: "EGFR unk"},
    "tp53_alteration":   {0: "TP53 WT", 1: "TP53 alt", 2: "TP53 unk"},
    "pten":              {0: "PTEN WT", 1: "PTEN mut", 2: "PTEN unk"},
    "atrx":              {0: "ATRX WT", 1: "ATRX loss", 2: "ATRX unk"},
    "cdkn2ab_deletion":  {0: "CDKN2A/B intact", 1: "CDKN2A/B del", 2: "CDKN2A/B unk"},
    "tert_promoter":     {0: "TERT WT", 1: "TERT mut", 2: "TERT unk"},
    "chr7_gain_chr10_loss": {0: "chr7/10 normal", 1: "+7/-10", 2: "+7/-10 unk"},
    "codeletion_1p19q_detail": {"0": "1p/19q intact", "1": "1p/19q codeleted"},
}

KEY_GENO = ["idh1", "mgmt_methylation", "egfr_amp", "tp53_alteration",
            "cdkn2ab_deletion", "tert_promoter", "codeletion_1p19q_detail"]


# ═══════════════════════════════════════════════════════════════════════════════
# 1. 数据加载 & Case 选择
# ═══════════════════════════════════════════════════════════════════════════════

def load_clinical():
    with open(CLINICAL_JSON) as f:
        return json.load(f)["patients"]


def tp_num(tp_id: str) -> int:
    m = re.search(r"(\d+)", str(tp_id))
    return int(m.group(1)) if m else -1


def find_tp1_tp2_pairs(patients: dict) -> list:
    """返回所有 TP1->TP2（或首对相邻 TP）有 MRI 文件的有效 case。"""
    pairs = []
    for pid, p in patients.items():
        tl = sorted(p.get("timeline", []), key=lambda x: tp_num(x["tp_id"]))
        if len(tl) < 2:
            continue
        for i in range(len(tl) - 1):
            pre_tp = tl[i]
            post_tp = tl[i + 1]
            pre_id  = f"Timepoint_{tp_num(pre_tp['tp_id'])}"
            post_id = f"Timepoint_{tp_num(post_tp['tp_id'])}"
            # 检查 MRI 文件是否存在
            mri_dir_pre = MRI_ROOT / pid / pre_id
            if not (mri_dir_pre / f"{pid}_{pre_id}_brain_t1c.nii.gz").exists():
                continue
            # 检查 post actions 非空
            post_actions = post_tp.get("actions", {})
            if not any(post_actions.values()):
                continue
            survival = post_tp.get("survival", {})
            if float(survival.get("survival_from_tp_days", -1)) < 0:
                continue
            pairs.append({
                "pid": pid,
                "context_static": p["context_static"],
                "pre_tp": pre_tp,
                "post_tp": post_tp,
                "pre_id": pre_id,
                "post_id": post_id,
                "mri_dir": mri_dir_pre,
                "survival_days": float(survival["survival_from_tp_days"]),
                "event": int(survival.get("event_indicator", 0)),
                "interval_days": post_tp.get("mri_day", 0) - pre_tp.get("mri_day", 0),
            })
    return pairs


def select_diverse_cases(pairs: list, n: int = 20, seed: int = RANDOM_SEED) -> list:
    """按 IDH 状态、MGMT、grade 分层，随机抽取 n 个多样化 case。"""
    rng = random.Random(seed)

    def stratum(p):
        g = p["context_static"].get("genomics", {})
        idh = g.get("idh1", 2)
        mgmt = g.get("mgmt_methylation", 2)
        grade = int(p["context_static"].get("who_grade", 4))
        return (idh, mgmt, grade)

    # 过滤：只保留有 TP1 的
    tp1_pairs = [p for p in pairs if tp_num(p["pre_tp"]["tp_id"]) == 1]
    # 若 TP1 不够 20，允许其他首 TP
    if len(tp1_pairs) < n:
        tp1_pairs = pairs

    # 分层
    strata: dict = {}
    for p in tp1_pairs:
        k = stratum(p)
        strata.setdefault(k, []).append(p)

    selected = []
    keys = list(strata.keys())
    rng.shuffle(keys)
    # 每个 stratum 取 1 个直到满足 n
    for k in keys:
        if len(selected) >= n:
            break
        candidates = strata[k]
        rng.shuffle(candidates)
        selected.append(candidates[0])
    # 不足 n：从剩余里补
    used_pids = {p["pid"] for p in selected}
    remaining = [p for p in tp1_pairs if p["pid"] not in used_pids]
    rng.shuffle(remaining)
    selected.extend(remaining[:n - len(selected)])
    return selected[:n]


# ═══════════════════════════════════════════════════════════════════════════════
# 2. MRI 切片生成
# ═══════════════════════════════════════════════════════════════════════════════

def load_nifti_volume(path: Path) -> np.ndarray:
    img = nib.load(str(path))
    vol = img.get_fdata()
    # 确保方向一致：RAS
    vol = np.nan_to_num(vol, nan=0.0)
    return vol


def percentile_clip(vol: np.ndarray, lo=1, hi=99) -> np.ndarray:
    p_lo, p_hi = np.percentile(vol[vol > 0], [lo, hi]) if vol.max() > 0 else (0, 1)
    return np.clip((vol - p_lo) / max(p_hi - p_lo, 1e-6), 0, 1)


def find_tumor_center(mask_vol: np.ndarray) -> tuple:
    """从 tumorMask 找肿瘤中心切片坐标。"""
    if mask_vol.max() == 0:
        z = mask_vol.shape[2] // 2
        y = mask_vol.shape[1] // 2
        x = mask_vol.shape[0] // 2
    else:
        coords = np.argwhere(mask_vol > 0)
        center = coords.mean(axis=0).astype(int)
        x, y, z = center[0], center[1], center[2]
    return x, y, z


def generate_mri_panel(mri_dir: Path, pid: str, tp_id: str, out_path: Path):
    """生成 2×2 多模态 MRI 面板（轴位 / 矢状 / 冠状切片），保存为 PNG。"""
    modalities = {
        "T1c": f"{pid}_{tp_id}_brain_t1c.nii.gz",
        "T2w": f"{pid}_{tp_id}_brain_t2w.nii.gz",
        "T1n": f"{pid}_{tp_id}_brain_t1n.nii.gz",
        "T2f": f"{pid}_{tp_id}_brain_t2f.nii.gz",
    }
    mask_path = mri_dir / f"{pid}_{tp_id}_tumorMask.nii.gz"

    vols = {}
    for name, fname in modalities.items():
        p = mri_dir / fname
        if p.exists():
            vols[name] = percentile_clip(load_nifti_volume(p))

    if not vols:
        return False  # 无数据

    # 用任意模态找肿瘤中心
    ref_key = next(iter(vols))
    ref_vol = vols[ref_key]
    mask_vol = load_nifti_volume(mask_path) if mask_path.exists() else np.zeros_like(ref_vol)
    cx, cy, cz = find_tumor_center(mask_vol)

    fig, axes = plt.subplots(2, 2, figsize=(10, 10), facecolor="black")
    fig.suptitle(f"{pid} | {tp_id}", color="white", fontsize=14, y=0.02)

    for ax, (mod_name, vol) in zip(axes.flat, list(vols.items())[:4]):
        # 轴位切片（冠状平面），y 固定
        slc = vol[:, :, cz].T
        ax.imshow(slc, cmap="gray", origin="lower", interpolation="bilinear")
        # 叠加肿瘤 mask 轮廓
        if mask_vol.max() > 0:
            mask_slc = mask_vol[:, :, cz].T
            ax.contour(mask_slc, levels=[0.5], colors=["red"], linewidths=0.8, alpha=0.7)
        ax.set_title(mod_name, color="white", fontsize=12, pad=4)
        ax.axis("off")

    plt.tight_layout(pad=0.5)
    plt.savefig(out_path, dpi=120, bbox_inches="tight",
                facecolor="black", edgecolor="none")
    plt.close()
    return True


# ═══════════════════════════════════════════════════════════════════════════════
# 3. 治疗方案文字格式化
# ═══════════════════════════════════════════════════════════════════════════════

def _infer_cycles(item: dict) -> str:
    """推算 num_cycles：优先用 num_cycles，否则从 interval/cycle_length 推算，否则用时长估算。"""
    if item.get("num_cycles") is not None:
        return str(int(item["num_cycles"]))
    i_start = item.get("interval_start_day")
    i_end   = item.get("interval_end_day")
    cld     = item.get("cycle_length_days")
    if i_start is not None and i_end is not None and cld and cld > 0:
        n = max(1, round((i_end - i_start) / cld))
        return str(n)
    start = item.get("start_day")
    end   = item.get("end_day")
    if start is not None and end is not None and cld and cld > 0:
        n = max(1, round((end - start) / cld))
        return str(n)
    return "ongoing"


def fmt_actions(actions: dict) -> str:
    """将 actions dict 转为可读的治疗描述。"""
    if not actions or not any(v for v in actions.values()):
        return "  - Observation / No active therapy -"
    lines = []
    label_map = {
        "radiation":     "RT",
        "chemotherapy":  "Chemo",
        "immunotherapy": "Immuno",
        "additional_1":  "Adjuvant-1",
        "additional_2":  "Adjuvant-2",
        "other_therapy": "Other",
        "other":         "Other",
        "brachy":        "Brachy",
    }
    for category, items in actions.items():
        if not items:
            continue
        cat_label = label_map.get(category, category.replace("_", " ").title())
        for item in items:
            if category == "radiation":
                dose = item.get("dose_gy", "?")
                frac = item.get("fractions", "?")
                lines.append(f"  {cat_label}: {dose} Gy / {frac} fractions")
            elif category == "brachy":
                agent = item.get("agent", "unknown")
                lines.append(f"  {cat_label}: {agent}")
            else:
                agent  = item.get("agent", "unknown agent")
                cld    = item.get("cycle_length_days")
                cycles = _infer_cycles(item)
                if cld:
                    lines.append(f"  {cat_label}: {agent} x{cycles} cycles (q{int(cld)}d)")
                else:
                    lines.append(f"  {cat_label}: {agent} (continuous)")
    return "\n".join(lines) if lines else "  - No structured therapy data -"


def fmt_genomics(genomics: dict) -> str:
    """摘要关键基因组学，只展示 known（非 unknown）结果。"""
    parts = []
    for key in KEY_GENO:
        val = genomics.get(key)
        if val is None:
            continue
        label_dict = GENO_LABELS.get(key, {})
        label = label_dict.get(val, label_dict.get(str(val), "unk"))
        if "unk" not in label:
            parts.append(label)
    return "  " + ";  ".join(parts) if parts else "  Not fully characterized"


def ascii_safe(s: str) -> str:
    """Replace non-latin-1 characters for FPDF core fonts."""
    return (s.replace("-", "-").replace("-", "-")
             .replace("'", "'").replace("'", "'")
             .replace("“", '"').replace("”", '"')
             .replace("é", "e").replace("à", "a"))


def fmt_clinical_summary(case: dict, case_num: int) -> dict:
    cs = case["context_static"]
    pre_tp = case["pre_tp"]
    genomics = cs.get("genomics", {})
    diag = cs.get("primary_diagnosis", "glioma").upper()
    grade = int(cs.get("who_grade", 4))
    age = int(cs.get("age_at_diagnosis_years", 0))
    sex = cs.get("sex_at_birth", "unknown")
    race = cs.get("race", "unknown")
    days_dx = pre_tp.get("state", {}).get("days_since_dx", pre_tp.get("mri_day", 0))
    progression = pre_tp.get("state", {}).get("progression", {})
    prog_str = "Progression documented" if progression.get("occurred_up_to_tp") else "No prior progression"

    return {
        "case_num": case_num,
        "pid": case["pid"],
        "pre_tp_id": case["pre_tp"]["tp_id"],
        "post_tp_id": case["post_tp"]["tp_id"],
        "age": age,
        "sex": sex.capitalize(),
        "race": race.capitalize(),
        "diagnosis": diag,
        "who_grade": grade,
        "days_from_dx": days_dx,
        "genomics_str": fmt_genomics(genomics),
        "progression": prog_str,
        "interval_days": case["interval_days"],
        "survival_days": int(case["survival_days"]),
        "event": "Deceased" if case["event"] == 1 else "Alive / Censored",
        "gt_actions_str": fmt_actions(case["post_tp"].get("actions", {})),
        "pre_actions_str": fmt_actions(case["pre_tp"].get("actions", {})),
    }


# ═══════════════════════════════════════════════════════════════════════════════
# 4. 随机 A/B 分配
# ═══════════════════════════════════════════════════════════════════════════════

def assign_ab(cases_meta: list, clarity_plans: dict, seed: int = RANDOM_SEED) -> list:
    """
    对每个 case 随机决定 A=CLARITY/B=GT 或 A=GT/B=CLARITY。
    clarity_plans: {pid: str}  -  CLARITY 推荐文字（若无则占位）
    返回包含 ab_mapping 的 list。
    """
    rng = random.Random(seed + 1)
    result = []
    for i, m in enumerate(cases_meta):
        clarity_text = clarity_plans.get(m["pid"], "[CLARITY RECOMMENDATION PENDING]")
        gt_text = m["gt_actions_str"]
        swap = rng.random() < 0.5
        if swap:
            option_a, option_b = clarity_text, gt_text
            a_is = "CLARITY"
        else:
            option_a, option_b = gt_text, clarity_text
            a_is = "GT"
        result.append({**m,
            "option_a": option_a,
            "option_b": option_b,
            "a_is": a_is,
        })
    return result


# ═══════════════════════════════════════════════════════════════════════════════
# 5. PDF 生成（每个 case 一页）
# ═══════════════════════════════════════════════════════════════════════════════

class CasePDF(FPDF):
    def header(self):
        self.set_font("Helvetica", "B", 10)
        self.set_fill_color(30, 60, 120)
        self.set_text_color(255, 255, 255)
        self.cell(0, 8, "CLARITY vs GT - Blind Expert Evaluation  |  CONFIDENTIAL", ln=True, fill=True, align="C")
        self.set_text_color(0, 0, 0)
        self.ln(2)

    def footer(self):
        self.set_y(-12)
        self.set_font("Helvetica", "I", 8)
        self.set_text_color(100, 100, 100)
        self.cell(0, 5, f"Page {self.page_no()} - Do not discuss with other reviewers", align="C")
        self.set_text_color(0, 0, 0)

    def section_title(self, txt: str):
        self.set_font("Helvetica", "B", 10)
        self.set_fill_color(220, 230, 245)
        self.cell(0, 7, ascii_safe(f"  {txt}"), ln=True, fill=True)
        self.ln(1)

    def body_line(self, label: str, value: str):
        label_w = 48
        val_w   = self.w - self.l_margin - self.r_margin - label_w
        y0 = self.get_y()
        self.set_font("Helvetica", "B", 9)
        self.cell(label_w, 6, ascii_safe(f"  {label}:"), border=0)
        x_val = self.l_margin + label_w
        self.set_xy(x_val, y0)
        self.set_font("Helvetica", "", 9)
        self.multi_cell(val_w, 6, ascii_safe(value))

    def option_box(self, letter: str, content: str, color: tuple):
        r, g, b = color
        self.set_fill_color(r, g, b)
        self.set_font("Helvetica", "B", 11)
        self.cell(0, 8, ascii_safe(f"  Treatment Option {letter}"), ln=True, fill=True)
        self.set_font("Courier", "", 9)
        self.set_fill_color(245, 245, 250)
        self.multi_cell(0, 5.5, ascii_safe(content), fill=True, border=1)
        self.ln(2)

    def judgment_box(self):
        self.section_title("Expert Judgment (select ONE)")
        self.set_font("Helvetica", "", 10)
        for label in [
            "[ ]  Option A is more clinically reasonable",
            "[ ]  Option B is more clinically reasonable",
            "[ ]  Both options are clinically equivalent / acceptable",
        ]:
            self.cell(0, 8, f"     {label}", ln=True)
        self.ln(3)
        self.set_font("Helvetica", "B", 9)
        self.cell(0, 6, "  Optional comment:", ln=True)
        self.set_fill_color(255, 255, 200)
        self.cell(0, 20, "", border=1, fill=True, ln=True)
        self.ln(2)

    def confidence_box(self):
        self.section_title("Confidence Level")
        self.set_font("Helvetica", "", 9)
        self.cell(0, 6, "  (1=Very uncertain  2=Uncertain  3=Moderate  4=Confident  5=Very confident)", ln=True)
        self.set_font("Helvetica", "", 10)
        self.cell(0, 8, "     [ ] 1     [ ] 2     [ ] 3     [ ] 4     [ ] 5", ln=True)
        self.ln(4)


def add_case_to_pdf(pdf: "CasePDF", case_ab: dict, mri_png: Path):
    """Add one case as a new page (or pages) to an existing CasePDF object."""
    pdf.add_page()

    # ── Case header ──────────────────────────────────────────────────────────
    pdf.set_font("Helvetica", "B", 14)
    pdf.set_text_color(30, 60, 120)
    pdf.cell(0, 10,
        f"Case {case_ab['case_num']:02d}   |   "
        f"{case_ab['pre_tp_id']} -> {case_ab['post_tp_id']}",
        ln=True)
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Helvetica", "I", 8)
    pdf.cell(0, 5, "Patient identity has been anonymized. Do not attempt to de-identify.", ln=True)
    pdf.ln(2)

    # ── Patient Summary ──────────────────────────────────────────────────────
    pdf.section_title("Patient Summary")
    pdf.body_line("Age / Sex", f"{case_ab['age']} / {case_ab['sex']}")
    pdf.body_line("Race", case_ab["race"])
    pdf.body_line("Diagnosis", f"{case_ab['diagnosis']}, WHO Grade {case_ab['who_grade']}")
    pdf.body_line("Days from Dx to this scan", str(case_ab["days_from_dx"]))
    pdf.body_line("Prior treatment (pre-scan)", case_ab["pre_actions_str"])
    pdf.body_line("Disease status", case_ab["progression"])
    pdf.body_line("Key molecular markers", case_ab["genomics_str"])
    pdf.body_line("Interval to next scan", f"{case_ab['interval_days']} days")
    pdf.ln(2)

    # ── MRI Panel ────────────────────────────────────────────────────────────
    pdf.section_title("MRI at Baseline (pre-treatment scan)")
    if mri_png.exists():
        # 动态缩放：A4 宽约 190mm，保持正方比
        img_w = min(180, pdf.w - 2 * pdf.l_margin)
        pdf.image(str(mri_png), x=pdf.l_margin, w=img_w)
        pdf.set_font("Helvetica", "I", 8)
        pdf.cell(0, 5,
            "Red contour = tumor mask.  Columns: T1c, T2w, T1n, T2f  (axial slice through tumor center)",
            ln=True, align="C")
    else:
        pdf.set_font("Helvetica", "I", 9)
        pdf.cell(0, 8, "  [MRI image not available for this case]", ln=True)
    pdf.ln(3)

    # ── Treatment Options ────────────────────────────────────────────────────
    pdf.section_title("Proposed Treatment Plans for Next Interval")
    pdf.set_font("Helvetica", "I", 8)
    pdf.cell(0, 5,
        "One option is a real physician-prescribed regimen; the other is an AI-generated candidate. "
        "You do not know which is which.",
        ln=True)
    pdf.ln(2)
    pdf.option_box("A", case_ab["option_a"], (200, 220, 255))
    pdf.option_box("B", case_ab["option_b"], (200, 255, 210))

    # ── Judgment ─────────────────────────────────────────────────────────────
    pdf.judgment_box()
    pdf.confidence_box()

    pdf.set_font("Helvetica", "I", 8)
    pdf.set_text_color(120, 120, 120)
    pdf.cell(0, 5, f"Generated: {date.today()}  |  Study ID: CLARITY-BLIND-{case_ab['case_num']:02d}", ln=True, align="R")
    pdf.set_text_color(0, 0, 0)


def build_all_cases_pdf(cases_ab: list, mri_pngs: list, out_path: Path):
    """Build one PDF containing all cases, one case per page-group."""
    pdf = CasePDF()
    pdf.set_auto_page_break(True, margin=12)
    for case_ab, mri_png in zip(cases_ab, mri_pngs):
        add_case_to_pdf(pdf, case_ab, mri_png)
    pdf.output(str(out_path))


# ═══════════════════════════════════════════════════════════════════════════════
# 6. Excel 输出
# ═══════════════════════════════════════════════════════════════════════════════

HEADER_FILL   = PatternFill("solid", fgColor="1E3C78")
HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Calibri")
ALT_FILL      = PatternFill("solid", fgColor="EEF2FF")
CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT          = Alignment(horizontal="left",   vertical="center", wrap_text=True)
THIN_BORDER   = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)


def style_header_row(ws, row=1):
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def style_data_row(ws, row_idx: int):
    fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()
    for cell in ws[row_idx]:
        cell.fill = fill
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def build_expert_sheet(cases_ab: list, out_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expert Response"

    headers = [
        "Case No.", "Expert Name", "Date",
        "Choice\n(A / B / Equivalent)", "Confidence\n(1-5)",
        "Comment (optional)"
    ]
    ws.append(headers)
    style_header_row(ws)

    for i, c in enumerate(cases_ab, 2):
        ws.append([c["case_num"], "", "", "", "", ""])
        style_data_row(ws, i)

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 40
    ws.row_dimensions[1].height = 40

    # Instructions sheet
    info = wb.create_sheet("Instructions")
    instructions = [
        ("CLARITY vs GT Blind Expert Evaluation", None),
        ("", None),
        ("Purpose:", "Evaluate two candidate treatment regimens for each case and indicate which is more clinically reasonable."),
        ("Blinding:", "You do NOT know which option (A or B) comes from a real physician and which is AI-generated. Please evaluate on clinical merit only."),
        ("Choices:", "A / B / Equivalent  - 'Equivalent' means both options are clinically acceptable."),
        ("Confidence:", "Rate your confidence: 1=Very uncertain ... 5=Very confident."),
        ("Independence:", "Please do NOT discuss cases with other reviewers until you have submitted your responses."),
        ("Questions:", "Contact the study coordinator if you need clarification."),
        ("", None),
        ("Response deadline:", "_______________"),
        ("Your name:", "_______________"),
    ]
    for row in instructions:
        info.append(row)
    for r in range(1, len(instructions) + 1):
        for c in range(1, 3):
            info.cell(r, c).alignment = LEFT

    info.column_dimensions["A"].width = 20
    info.column_dimensions["B"].width = 80
    info["A1"].font = Font(bold=True, size=14)

    wb.save(out_path)


def build_researcher_mapping(cases_ab: list, out_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GT Mapping (CONFIDENTIAL)"

    headers = ["Case No.", "Patient ID", "Pre TP", "Post TP",
               "A is", "B is", "GT Treatment", "CLARITY Recommendation",
               "Survival Days", "Event"]
    ws.append(headers)
    style_header_row(ws)

    for i, c in enumerate(cases_ab, 2):
        ws.append([
            c["case_num"], c["pid"], c["pre_tp_id"], c["post_tp_id"],
            c["a_is"], "GT" if c["a_is"] == "CLARITY" else "CLARITY",
            c["gt_actions_str"], c["option_a"] if c["a_is"] == "CLARITY" else c["option_b"],
            c["survival_days"], c["event"],
        ])
        style_data_row(ws, i)

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 40
    ws.column_dimensions["H"].width = 40
    ws.column_dimensions["I"].width = 16
    ws.column_dimensions["J"].width = 12

    warn = wb.create_sheet("⚠ WARNING")
    warn["A1"] = "DO NOT SHARE WITH EXPERTS"
    warn["A1"].font = Font(bold=True, size=16, color="CC0000")
    warn["A2"] = "This sheet contains the A/B identity mapping. Keep strictly confidential until evaluation is complete."

    wb.save(out_path)


# ═══════════════════════════════════════════════════════════════════════════════
# 7. 主流程
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--clarity_json", type=str, default=None,
        help="Path to JSON: {pid: 'CLARITY recommendation text', ...}")
    parser.add_argument("--n_cases", type=int, default=NUM_CASES)
    args = parser.parse_args()

    clarity_plans: dict = {}
    if args.clarity_json:
        with open(args.clarity_json) as f:
            clarity_plans = json.load(f)
        print(f"[CLARITY] Loaded {len(clarity_plans)} recommendations from {args.clarity_json}")
    else:
        print("[CLARITY] No clarity_json provided - using placeholder text.")

    print("Loading clinical data...")
    patients = load_clinical()

    print("Finding valid TP->TP pairs with MRI data...")
    pairs = find_tp1_tp2_pairs(patients)
    print(f"  Found {len(pairs)} valid pairs")

    print(f"Selecting {args.n_cases} diverse cases...")
    selected = select_diverse_cases(pairs, n=args.n_cases)
    print(f"  Selected {len(selected)} cases")

    print("Building case metadata & clinical summaries...")
    cases_meta = [fmt_clinical_summary(c, i + 1) for i, c in enumerate(selected)]

    print("Assigning A/B randomization...")
    cases_ab = assign_ab(cases_meta, clarity_plans)

    print("\nGenerating MRI slices...")
    mri_pngs = []
    for i, raw in enumerate(selected):
        case_no = f"{i+1:02d}"
        print(f"  Case {case_no}: {raw['pid']} | {raw['pre_id']} -> {raw['post_id']}")
        mri_out = MRI_DIR / f"case_{case_no}_mri.png"
        ok = generate_mri_panel(raw["mri_dir"], raw["pid"], raw["pre_id"], mri_out)
        if not ok:
            print(f"    ⚠ MRI panel failed for {raw['pid']}")
        mri_pngs.append(mri_out)

    print("\nBuilding combined expert PDF (all cases)...")
    pdf_out = OUT_DIR / "expert_evaluation_cases.pdf"
    build_all_cases_pdf(cases_ab, mri_pngs, pdf_out)
    print(f"  ✓ {pdf_out.name}  ({len(cases_ab)} cases)")

    print("\nGenerating expert response sheet...")
    sheet_path = OUT_DIR / "expert_response_sheet.xlsx"
    build_expert_sheet(cases_ab, sheet_path)
    print(f"  ✓ {sheet_path.name}")

    print("Generating researcher mapping (CONFIDENTIAL)...")
    mapping_path = OUT_DIR / "researcher_mapping_CONFIDENTIAL.xlsx"
    build_researcher_mapping(cases_ab, mapping_path)
    print(f"  ✓ {mapping_path.name}")

    print("\n" + "="*60)
    print("DONE.")
    print(f"  Expert PDF:   {pdf_out}")
    print(f"  MRI slices:   {MRI_DIR}/")
    print(f"  Expert sheet: {sheet_path}")
    print(f"  GT mapping:   {mapping_path}  <- DO NOT SHARE WITH EXPERTS")
    print("="*60)


if __name__ == "__main__":
    main()
