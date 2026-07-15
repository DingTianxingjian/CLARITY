"""
CLARITY vs GT Blind Expert Evaluation — LaTeX Generator
========================================================
Output:
  expert_evaluation_cases.tex   -- 给专家的盲态 case 材料（含 GT，CLARITY 留空）
  mri_slices/case_XX_mri.png    -- MRI 切片图
  expert_response_sheet.xlsx    -- 专家填写表
  researcher_mapping_CONFIDENTIAL.xlsx  -- 研究者保存

用法：
  python generate_latex.py [--clarity_json PATH]
  pdflatex expert_evaluation_cases.tex   # 由用户自行渲染
"""

import os, sys, json, random, re, argparse, textwrap
import numpy as np
import nibabel as nib
import matplotlib; matplotlib.use("Agg")
import matplotlib.pyplot as plt
from pathlib import Path
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT          = Path(__file__).resolve().parents[2]
MRI_ROOT      = ROOT / "datasets" / "MU-Glioma-Post"
CLINICAL_JSON = ROOT / "Predictor" / "dataset" / "MU_Glioma_Post" / "clinical_latest.json"
OUT_DIR       = Path(__file__).parent
MRI_DIR       = OUT_DIR / "mri_slices"
MRI_DIR.mkdir(exist_ok=True)

NUM_CASES   = 20
RANDOM_SEED = 2024

# ── Claude (AI) treatment plans — LaTeX lines matching fmt_actions() format ──
# Format identical to GT: \textbf{Label}: Agent $\times$N cycles (qXd) | (continuous)
CLAUDE_PLANS = {
    # 26F salvage IDH-WT astrocytoma — Bev + TMZ rechallenge
    "PatientID_0210": [
        r"\textbf{Immuno}: Bevacizumab $\times$ongoing cycles (q14d)",
        r"\textbf{Adjuvant}: Temozolomide $\times$6 cycles (q28d)",
    ],
    # 35M newly diagnosed IDH-mut grade 4 — full Stupp + Optune
    "PatientID_0207": [
        r"\textbf{RT}: 60.0\,Gy / 30 fractions",
        r"\textbf{Chemo}: Temozolomide (continuous)",
        r"\textbf{Adjuvant}: Temozolomide $\times$6 cycles (q28d)",
        r"\textbf{Other}: Optune TTF (continuous)",
    ],
    # 45M IDH-mut grade 4 MGMT-met — Stupp + extended adjuvant TMZ + Optune
    "PatientID_0209": [
        r"\textbf{RT}: 60.0\,Gy / 30 fractions",
        r"\textbf{Chemo}: Temozolomide (continuous)",
        r"\textbf{Adjuvant}: Temozolomide $\times$12 cycles (q28d)",
        r"\textbf{Other}: Optune TTF (continuous)",
    ],
    # 63F IDH-WT grade 3 MGMT-met — CATNON: RT 59.4Gy + TMZ concurrent + 12 adj
    "PatientID_0188": [
        r"\textbf{RT}: 59.4\,Gy / 33 fractions",
        r"\textbf{Chemo}: Temozolomide (continuous)",
        r"\textbf{Adjuvant}: Temozolomide $\times$12 cycles (q28d)",
    ],
    # 28F oligodendroglioma grade 2 1p/19q codeleted — continue PCV per RTOG 9802
    "PatientID_0242": [
        r"\textbf{Adjuvant}: PCV $\times$6 cycles (q42d)",
    ],
    # 55F GBM IDH-WT post-CRT — standard adjuvant TMZ + Optune
    "PatientID_0036": [
        r"\textbf{Adjuvant}: Temozolomide $\times$6 cycles (q28d)",
        r"\textbf{Other}: Optune TTF (continuous)",
    ],
    # 25M astrocytoma grade 2 deep eloquent — RT 50.4 Gy + adjuvant TMZ
    "PatientID_0236": [
        r"\textbf{RT}: 50.4\,Gy / 28 fractions",
        r"\textbf{Adjuvant}: Temozolomide $\times$6 cycles (q28d)",
    ],
    # 38F IDH-mut grade 2 MGMT-unmeth — RT 54 Gy + concurrent + adjuvant TMZ
    "PatientID_0195": [
        r"\textbf{RT}: 54.0\,Gy / 30 fractions",
        r"\textbf{Chemo}: Temozolomide (continuous)",
        r"\textbf{Adjuvant}: Temozolomide $\times$6 cycles (q28d)",
    ],
    # 43F oligodendroglioma grade 2 1p/19q codeleted — adjuvant PCV per RTOG 9802
    "PatientID_0275": [
        r"\textbf{Adjuvant}: PCV $\times$6 cycles (q42d)",
    ],
    # 54M GBM IDH-WT MGMT-unmeth post-CRT — adjuvant TMZ + Optune
    "PatientID_0072": [
        r"\textbf{Adjuvant}: Temozolomide $\times$6 cycles (q28d)",
        r"\textbf{Other}: Optune TTF (continuous)",
    ],
    # 32F IDH-WT grade 2 with GBM-like imaging — treat as GBM: RT + TMZ + adj
    "PatientID_0235": [
        r"\textbf{RT}: 60.0\,Gy / 30 fractions",
        r"\textbf{Chemo}: Temozolomide (continuous)",
        r"\textbf{Adjuvant}: Temozolomide $\times$6 cycles (q28d)",
    ],
    # 53M GBM IDH-WT MGMT-met post-CRT — adjuvant TMZ + Optune
    "PatientID_0033": [
        r"\textbf{Adjuvant}: Temozolomide $\times$6 cycles (q28d)",
        r"\textbf{Other}: Optune TTF (continuous)",
    ],
    # 56F GBM grade 4 MGMT-unmeth post-CRT — adjuvant TMZ + Optune
    "PatientID_0272": [
        r"\textbf{Adjuvant}: Temozolomide $\times$6 cycles (q28d)",
        r"\textbf{Other}: Optune TTF (continuous)",
    ],
    # 37M GBM deep eloquent 2/6 adj TMZ done — continue remaining cycles + Optune
    "PatientID_0085": [
        r"\textbf{Adjuvant}: Temozolomide $\times$4 cycles (q28d)",
        r"\textbf{Other}: Optune TTF (continuous)",
    ],
    # 24F IDH-mut grade 4 MGMT-met — Stupp + extended 12-cycle adj TMZ + Optune
    "PatientID_0204": [
        r"\textbf{RT}: 60.0\,Gy / 30 fractions",
        r"\textbf{Chemo}: Temozolomide (continuous)",
        r"\textbf{Adjuvant}: Temozolomide $\times$12 cycles (q28d)",
        r"\textbf{Other}: Optune TTF (continuous)",
    ],
    # 59F recurrent GBM MGMT-met rapid progression — BELOB: Bev + Lomustine
    "PatientID_0250": [
        r"\textbf{Immuno}: Bevacizumab $\times$ongoing cycles (q14d)",
        r"\textbf{Adjuvant}: Lomustine (CCNU) $\times$ongoing cycles (q42d)",
    ],
    # 65F GBM IDH-WT MGMT-met rapid progression — Bev + Lomustine salvage
    "PatientID_0054": [
        r"\textbf{Immuno}: Bevacizumab $\times$ongoing cycles (q14d)",
        r"\textbf{Adjuvant}: Lomustine $\times$ongoing cycles (q42d)",
    ],
    # 66M GBM MGMT-met EGFR-amp post-CRT — adjuvant TMZ + Optune (TMZ mandated)
    "PatientID_0267": [
        r"\textbf{Adjuvant}: Temozolomide $\times$6 cycles (q28d)",
        r"\textbf{Other}: Optune TTF (continuous)",
    ],
    # 53M GBM IDH-WT MGMT-unmeth Gliadel post-CRT — adjuvant TMZ + Optune
    "PatientID_0038": [
        r"\textbf{Adjuvant}: Temozolomide $\times$6 cycles (q28d)",
        r"\textbf{Other}: Optune TTF (continuous)",
    ],
    # 61F GBM massive IDH-WT MGMT-unmeth rapid progression — Bev + supportive care
    "PatientID_0087": [
        r"\textbf{Immuno}: Bevacizumab $\times$ongoing cycles (q14d)",
        r"\textbf{Other}: Dexamethasone (continuous)",
    ],
}

GENO_LABELS = {
    "idh1":             {0: "IDH1 WT", 1: "IDH1 mut", 2: None},
    "idh2":             {0: "IDH2 WT", 1: "IDH2 mut", 2: None},
    "mgmt_methylation": {0: "MGMT unmethylated", 1: "MGMT methylated", 2: None},
    "egfr_amp":         {0: None, 1: "EGFR amp", 2: None},
    "tp53_alteration":  {0: None, 1: "TP53 alt", 2: None},
    "pten":             {0: None, 1: "PTEN mut", 2: None},
    "atrx":             {0: None, 1: "ATRX loss", 2: None},
    "cdkn2ab_deletion": {0: None, 1: "CDKN2A/B del", 2: None},
    "tert_promoter":    {0: None, 1: "TERT mut", 2: None},
    "chr7_gain_chr10_loss": {0: None, 1: "+7/-10", 2: None},
    "codeletion_1p19q_detail": {"0": None, "1": "1p/19q codeleted"},
}
KEY_GENO = ["idh1", "idh2", "mgmt_methylation", "egfr_amp", "tp53_alteration",
            "cdkn2ab_deletion", "tert_promoter", "codeletion_1p19q_detail"]


# ── helpers ──────────────────────────────────────────────────────────────────

def tex(s: str) -> str:
    """Escape special LaTeX characters."""
    s = str(s)
    for ch, rep in [("&","\\&"),("%","\\%"),("$","\\$"),("#","\\#"),
                    ("_","\\_"),("{","\\{"),("}","\\}"),("~","\\textasciitilde{}"),
                    ("^","\\textasciicircum{}"),("\\","\\textbackslash{}")]:
        s = s.replace(ch, rep)
    return s


def load_clinical():
    with open(CLINICAL_JSON) as f:
        return json.load(f)["patients"]


def tp_num(tp_id):
    m = re.search(r"(\d+)", str(tp_id))
    return int(m.group(1)) if m else -1


def find_pairs(patients):
    pairs = []
    for pid, p in patients.items():
        tl = sorted(p.get("timeline", []), key=lambda x: tp_num(x["tp_id"]))
        for i in range(len(tl) - 1):
            pre, post = tl[i], tl[i+1]
            pre_id  = f"Timepoint_{tp_num(pre['tp_id'])}"
            post_id = f"Timepoint_{tp_num(post['tp_id'])}"
            mri_dir = MRI_ROOT / pid / pre_id
            if not (mri_dir / f"{pid}_{pre_id}_brain_t1c.nii.gz").exists():
                continue
            if not any(post.get("actions", {}).values()):
                continue
            surv = post.get("survival", {})
            if float(surv.get("survival_from_tp_days", -1)) < 0:
                continue
            pairs.append({
                "pid": pid, "context_static": p["context_static"],
                "pre_tp": pre, "post_tp": post,
                "pre_id": pre_id, "post_id": post_id,
                "mri_dir": mri_dir,
                "survival_days": float(surv["survival_from_tp_days"]),
                "event": int(surv.get("event_indicator", 0)),
                "interval_days": post.get("mri_day", 0) - pre.get("mri_day", 0),
            })
    return pairs


def select_diverse(pairs, n=20, seed=RANDOM_SEED):
    rng = random.Random(seed)
    tp1 = [p for p in pairs if tp_num(p["pre_tp"]["tp_id"]) == 1]
    if len(tp1) < n:
        tp1 = pairs
    strata = {}
    for p in tp1:
        g = p["context_static"].get("genomics", {})
        k = (g.get("idh1", 2), g.get("mgmt_methylation", 2),
             int(p["context_static"].get("who_grade", 4)))
        strata.setdefault(k, []).append(p)
    selected, used = [], set()
    for k in sorted(strata, key=lambda x: rng.random()):
        if len(selected) >= n: break
        cands = [p for p in strata[k] if p["pid"] not in used]
        if cands:
            c = rng.choice(cands)
            selected.append(c); used.add(c["pid"])
    rest = [p for p in tp1 if p["pid"] not in used]
    rng.shuffle(rest)
    selected.extend(rest[:n - len(selected)])
    return selected[:n]


# ── MRI slices ───────────────────────────────────────────────────────────────

def load_clip(path):
    vol = nib.load(str(path)).get_fdata()
    vol = np.nan_to_num(vol)
    p1, p99 = (np.percentile(vol[vol > 0], [1, 99])
                if vol.max() > 0 else (0, 1))
    return np.clip((vol - p1) / max(p99 - p1, 1e-6), 0, 1)


def tumor_center(mask):
    if mask.max() == 0:
        return [s // 2 for s in mask.shape]
    return np.argwhere(mask > 0).mean(axis=0).astype(int).tolist()


def make_mri_panel(mri_dir, pid, tp_id, out_path):
    mods = {"T1c": "t1c", "T2w": "t2w", "T1n": "t1n", "T2f": "t2f"}
    vols = {}
    for label, suffix in mods.items():
        p = mri_dir / f"{pid}_{tp_id}_brain_{suffix}.nii.gz"
        if p.exists():
            vols[label] = load_clip(p)
    if not vols:
        return False
    mask_p = mri_dir / f"{pid}_{tp_id}_tumorMask.nii.gz"
    mask = nib.load(str(mask_p)).get_fdata() if mask_p.exists() else np.zeros((1,1,1))
    cx, cy, cz = tumor_center(mask)

    fig, axes = plt.subplots(1, 4, figsize=(16, 4), facecolor="black")
    for ax, (label, vol) in zip(axes, list(vols.items())[:4]):
        slc = vol[:, :, cz].T
        ax.imshow(slc, cmap="gray", origin="lower", interpolation="bilinear")
        if mask.max() > 0:
            ax.contour(mask[:, :, cz].T, levels=[0.5],
                       colors=["red"], linewidths=1.0, alpha=0.8)
        ax.set_title(label, color="white", fontsize=13, pad=3)
        ax.axis("off")
    plt.tight_layout(pad=0.3)
    plt.savefig(out_path, dpi=130, bbox_inches="tight",
                facecolor="black", edgecolor="none")
    plt.close()
    return True


# ── treatment formatting ─────────────────────────────────────────────────────

def infer_cycles(item):
    if item.get("num_cycles") is not None:
        return str(int(item["num_cycles"]))
    s, e = item.get("interval_start_day"), item.get("interval_end_day")
    cld = item.get("cycle_length_days")
    if s is not None and e is not None and cld and cld > 0:
        return str(max(1, round((e - s) / cld)))
    s2, e2 = item.get("start_day"), item.get("end_day")
    if s2 is not None and e2 is not None and cld and cld > 0:
        return str(max(1, round((e2 - s2) / cld)))
    return "ongoing"


def fmt_actions(actions):
    if not actions or not any(v for v in actions.values()):
        return ["\\textit{Observation / No active therapy}"]
    label_map = {
        "radiation": "RT", "chemotherapy": "Chemo",
        "immunotherapy": "Immuno", "additional_1": "Adjuvant",
        "additional_2": "Adjuvant-2", "other_therapy": "Other",
        "other": "Other", "brachy": "Brachy",
    }
    lines = []
    for cat, items in actions.items():
        if not items: continue
        label = label_map.get(cat, cat.replace("_", " ").title())
        for item in items:
            if cat == "radiation":
                dose = item.get("dose_gy", "?")
                frac = item.get("fractions", "?")
                lines.append(f"\\textbf{{{tex(label)}}}: {tex(dose)}\\,Gy / {tex(frac)} fractions")
            elif cat == "brachy":
                lines.append(f"\\textbf{{{tex(label)}}}: {tex(item.get('agent',''))}")
            else:
                agent = tex(item.get("agent", "unknown"))
                cld = item.get("cycle_length_days")
                cyc = infer_cycles(item)
                if cld:
                    lines.append(f"\\textbf{{{tex(label)}}}: {agent} $\\times${tex(cyc)} cycles (q{tex(int(cld))}d)")
                else:
                    lines.append(f"\\textbf{{{tex(label)}}}: {agent} (continuous)")
    return lines if lines else ["\\textit{No structured therapy data}"]


def fmt_genomics(genomics):
    parts = []
    for key in KEY_GENO:
        val = genomics.get(key)
        lmap = GENO_LABELS.get(key, {})
        label = lmap.get(val, lmap.get(str(val)))
        if label:
            parts.append(label)
    return tex(";  ".join(parts)) if parts else "\\textit{Not fully characterized}"


def fmt_case_meta(case, num):
    cs = case["context_static"]
    pre = case["pre_tp"]
    g = cs.get("genomics", {})
    days_dx = pre.get("state", {}).get("days_since_dx", pre.get("mri_day", 0))
    prog = pre.get("state", {}).get("progression", {})
    prog_str = "Progression documented" if prog.get("occurred_up_to_tp") else "No prior progression"
    return {
        "case_num": num,
        "pid": case["pid"],
        "pre_tp_id": pre["tp_id"],
        "post_tp_id": case["post_tp"]["tp_id"],
        "age": int(cs.get("age_at_diagnosis_years", 0)),
        "sex": cs.get("sex_at_birth", "unknown").capitalize(),
        "race": cs.get("race", "unknown").capitalize(),
        "diagnosis": cs.get("primary_diagnosis", "glioma").upper(),
        "who_grade": int(cs.get("who_grade", 4)),
        "days_from_dx": days_dx,
        "genomics_str": fmt_genomics(g),
        "progression": tex(prog_str),
        "interval_days": case["interval_days"],
        "gt_lines": fmt_actions(case["post_tp"].get("actions", {})),
        "pre_lines": fmt_actions(pre.get("actions", {})),
    }


# ── A/B randomization ────────────────────────────────────────────────────────

def assign_ab(cases_meta, clarity_plans, seed=RANDOM_SEED):
    rng = random.Random(seed + 1)
    result = []
    for m in cases_meta:
        clarity_lines = clarity_plans.get(m["pid"])
        gt_lines = m["gt_lines"]
        swap = rng.random() < 0.5
        if swap:
            a_lines, b_lines, a_is = clarity_lines or ["[CLARITY RECOMMENDATION PENDING]"], gt_lines, "CLARITY"
        else:
            a_lines, b_lines, a_is = gt_lines, clarity_lines or ["[CLARITY RECOMMENDATION PENDING]"], "GT"
        result.append({**m, "a_lines": a_lines, "b_lines": b_lines, "a_is": a_is})
    return result


# ── LaTeX generation ─────────────────────────────────────────────────────────

PREAMBLE = r"""
\documentclass[a4paper,11pt]{article}
\usepackage[a4paper, margin=2cm]{geometry}
\usepackage{graphicx}
\usepackage{xcolor}
\usepackage{booktabs}
\usepackage{array}
\usepackage{tabularx}
\usepackage{tcolorbox}
\usepackage{enumitem}
\usepackage{tikz}
\usepackage{fancyhdr}
\usepackage{fontenc}
\usepackage{inputenc}
\usepackage{microtype}
\usepackage{parskip}
\usepackage{multirow}
\usepackage{amsmath}
\usepackage{wasysym}

\tcbuselibrary{skins, breakable}

\definecolor{headerblue}{RGB}{30,60,120}
\definecolor{optionA}{RGB}{210,228,255}
\definecolor{optionB}{RGB}{210,245,220}
\definecolor{lightgray}{RGB}{245,245,245}
\definecolor{sectiongray}{RGB}{220,230,245}
\definecolor{warnred}{RGB}{180,30,30}

\pagestyle{fancy}
\fancyhf{}
\fancyhead[L]{\small\textcolor{headerblue}{\textbf{CLARITY vs GT --- Blind Expert Evaluation}}}
\fancyhead[R]{\small\textcolor{gray}{CONFIDENTIAL}}
\fancyfoot[C]{\small\textcolor{gray}{Page \thepage\ --- Do not discuss with other reviewers}}
\renewcommand{\headrulewidth}{0.4pt}

\newcommand{\checkbox}{\Square}
\newcommand{\sectionbar}[1]{%
  \vspace{4pt}
  {\colorbox{sectiongray}{\makebox[\dimexpr\linewidth-2\fboxsep\relax][l]{\strut\textbf{\small #1}}}}\vspace{3pt}
}
\newcommand{\inforow}[2]{%
  \makebox[5cm][l]{\textbf{\small #1:}} {\small #2}\\[2pt]
}

\begin{document}
"""

POSTAMBLE = r"\end{document}"


def case_block(c, mri_path):
    mri_rel = os.path.relpath(str(mri_path), str(OUT_DIR))
    lines = []
    lines.append(r"\newpage")
    lines.append("")

    # ── header ──
    lines.append(r"\begin{tcolorbox}[colback=headerblue,colframe=headerblue,arc=3pt]")
    lines.append(r"  \color{white}\large\bfseries")
    lines.append(f"  Case {c['case_num']:02d} \\hfill {tex(c['pre_tp_id'])} $\\rightarrow$ {tex(c['post_tp_id'])}")
    lines.append(r"\end{tcolorbox}")
    lines.append(r"{\footnotesize\textit{Patient identity anonymized. Do not attempt to de-identify.}}")
    lines.append("")

    # ── patient summary ──
    lines.append(r"\sectionbar{Patient Summary}")
    lines.append(r"\begin{tabular}{@{}p{5.5cm}p{10cm}@{}}")
    def row(label, val):
        return f"  \\textbf{{\\small {tex(label)}:}} & {{\\small {val}}} \\\\"
    lines.append(row("Age / Sex", f"{c['age']} / {tex(c['sex'])}"))
    lines.append(row("Race", tex(c['race'])))
    lines.append(row("Diagnosis", f"{tex(c['diagnosis'])}, WHO Grade {c['who_grade']}"))
    lines.append(row("Days from diagnosis", str(c['days_from_dx'])))
    lines.append(row("Disease status", c['progression']))
    lines.append(row("Key molecular markers", c['genomics_str']))
    lines.append(row("Interval to next scan", f"{c['interval_days']} days"))
    lines.append(r"\end{tabular}")
    lines.append("")

    # ── prior treatment ──
    lines.append(r"\sectionbar{Prior Treatment (at pre-scan timepoint)}")
    lines.append(r"\begin{itemize}[leftmargin=1.5em,itemsep=0pt,topsep=2pt]")
    for l in c["pre_lines"]:
        lines.append(f"  \\item {l}")
    lines.append(r"\end{itemize}")
    lines.append("")

    # ── MRI ──
    lines.append(r"\sectionbar{MRI at Pre-treatment Timepoint}")
    if Path(mri_path).exists():
        lines.append(r"\begin{center}")
        lines.append(f"  \\includegraphics[width=\\linewidth]{{{mri_rel}}}")
        lines.append(r"\end{center}")
        lines.append(r"{\footnotesize\centering\textit{Red contour = tumor mask. Left to right: T1c, T2w, T1n, T2f (axial slice through tumor center).}\\}")
    else:
        lines.append(r"\textit{[MRI image not available]}")
    lines.append("")

    # ── treatment options ──
    lines.append(r"\sectionbar{Proposed Treatment Plans for Next Interval}")
    lines.append(r"{\footnotesize\textit{One option is the actual physician-prescribed regimen; the other is an AI-generated candidate. You do not know which is which.}}")
    lines.append("")

    for letter, opt_lines, color in [("A", c["a_lines"], "optionA"),
                                      ("B", c["b_lines"], "optionB")]:
        lines.append(f"\\begin{{tcolorbox}}[colback={color},colframe={color}!60!black,arc=3pt,title={{\\textbf{{Treatment Option {letter}}}}},fonttitle=\\normalsize\\bfseries]")
        lines.append(r"  \begin{itemize}[leftmargin=1.5em,itemsep=1pt,topsep=0pt]")
        for l in opt_lines:
            lines.append(f"    \\item {l}")
        lines.append(r"  \end{itemize}")
        lines.append(r"\end{tcolorbox}")
        lines.append("")

    # ── judgment ──
    lines.append(r"\sectionbar{Expert Judgment (circle or mark ONE)}")
    lines.append(r"\begin{enumerate}[label=\alph*.,leftmargin=2em,itemsep=4pt,topsep=2pt]")
    lines.append(r"  \item[\checkbox] \textbf{Option A} is more clinically reasonable")
    lines.append(r"  \item[\checkbox] \textbf{Option B} is more clinically reasonable")
    lines.append(r"  \item[\checkbox] Both options are \textbf{clinically equivalent} / acceptable")
    lines.append(r"\end{enumerate}")
    lines.append("")

    # ── confidence ──
    lines.append(r"\sectionbar{Confidence Level \normalfont\small(1 = Very uncertain \quad 5 = Very confident)}")
    lines.append(r"\quad $\checkbox$~1 \qquad $\checkbox$~2 \qquad $\checkbox$~3 \qquad $\checkbox$~4 \qquad $\checkbox$~5")
    lines.append("")

    # ── comment ──
    lines.append(r"\sectionbar{Optional Comment}")
    lines.append(r"\vspace{1.5cm}\hrule\vspace{0.8cm}\hrule")
    lines.append("")

    # ── footer note ──
    lines.append(r"{\footnotesize\textcolor{gray}{\hfill Generated: "
                 + str(date.today())
                 + r" \quad Study ID: CLARITY-BLIND-"
                 + f"{c['case_num']:02d}"
                 + r"}}")
    return "\n".join(lines)


def build_latex(cases_ab, mri_pngs, out_path):
    parts = [PREAMBLE]
    for c, mri in zip(cases_ab, mri_pngs):
        parts.append(case_block(c, mri))
    parts.append(POSTAMBLE)
    out_path.write_text("\n".join(parts), encoding="utf-8")


# ── Excel outputs (unchanged from generate_cases.py) ─────────────────────────

def _style_header(ws):
    fill = PatternFill("solid", fgColor="1E3C78")
    font = Font(bold=True, color="FFFFFF", name="Calibri")
    for cell in ws[1]:
        cell.font = font; cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                             top=Side(style="thin"), bottom=Side(style="thin"))


def build_expert_sheet(cases_ab, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Expert Response"
    ws.append(["Case No.", "Expert Name", "Date",
               "Choice\n(A / B / Equivalent)", "Confidence\n(1-5)", "Comment (optional)"])
    _style_header(ws)
    for c in cases_ab:
        ws.append([c["case_num"], "", "", "", "", ""])
    for col, w in zip("ABCDEF", [10, 18, 14, 22, 16, 45]):
        ws.column_dimensions[col].width = w
    ws.row_dimensions[1].height = 36

    info = wb.create_sheet("Instructions")
    rows = [
        ("CLARITY vs GT -- Blind Expert Evaluation", ""),
        ("", ""),
        ("Purpose", "Evaluate two treatment regimens per case and select the more clinically reasonable one."),
        ("Blinding", "You do NOT know which option (A or B) is from a real physician and which is AI-generated."),
        ("Choices", "A / B / Equivalent  --  'Equivalent' means both options are clinically acceptable."),
        ("Confidence", "1 = Very uncertain   ...   5 = Very confident"),
        ("Independence", "Do NOT discuss cases with other reviewers until you have submitted your responses."),
        ("", ""),
        ("Your name", "_______________"),
        ("Response deadline", "_______________"),
    ]
    for r in rows:
        info.append(r)
    info["A1"].font = Font(bold=True, size=13)
    info.column_dimensions["A"].width = 20
    info.column_dimensions["B"].width = 80
    wb.save(out_path)


def build_mapping(cases_ab, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "GT Mapping (CONFIDENTIAL)"
    ws.append(["Case No.", "Patient ID", "Pre TP", "Post TP",
               "A is", "B is", "GT Treatment (lines)", "Survival Days", "Event"])
    _style_header(ws)
    for c in cases_ab:
        b_is = "GT" if c["a_is"] == "CLARITY" else "CLARITY"
        gt_text = " | ".join(c["gt_lines"])
        ws.append([c["case_num"], c["pid"], c["pre_tp_id"], c["post_tp_id"],
                   c["a_is"], b_is, gt_text, c.get("survival_days", ""), c.get("event", "")])
    for col, w in zip("ABCDEFGHI", [10,18,10,10,12,12,60,16,10]):
        ws.column_dimensions[col].width = w
    warn = wb.create_sheet("WARNING")
    warn["A1"] = "DO NOT SHARE WITH EXPERTS"
    warn["A1"].font = Font(bold=True, size=14, color="CC0000")
    wb.save(out_path)


# ── main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--clarity_json", default=None)
    parser.add_argument("--n_cases", type=int, default=NUM_CASES)
    args = parser.parse_args()

    if args.clarity_json:
        with open(args.clarity_json) as f:
            clarity_plans = json.load(f)
        print(f"[CLARITY] Loaded {len(clarity_plans)} plans from {args.clarity_json}.")
    else:
        clarity_plans = CLAUDE_PLANS
        print(f"[CLARITY] Using embedded Claude plans ({len(clarity_plans)} patients).")

    patients = load_clinical()
    pairs = find_pairs(patients)
    print(f"Found {len(pairs)} valid TP pairs")

    selected = select_diverse(pairs, n=args.n_cases)
    print(f"Selected {len(selected)} cases")

    cases_meta = [fmt_case_meta(c, i+1) for i, c in enumerate(selected)]
    cases_ab   = assign_ab(cases_meta, clarity_plans)

    print("Generating MRI panels...")
    mri_pngs = []
    for i, raw in enumerate(selected):
        out = MRI_DIR / f"case_{i+1:02d}_mri.png"
        ok  = make_mri_panel(raw["mri_dir"], raw["pid"], raw["pre_id"], out)
        mri_pngs.append(out)
        status = "ok" if ok else "MISSING"
        print(f"  {i+1:02d}: {raw['pid']} [{status}]")

    print("Writing LaTeX...")
    tex_out = OUT_DIR / "expert_evaluation_cases.tex"
    build_latex(cases_ab, mri_pngs, tex_out)
    print(f"  -> {tex_out}")

    build_expert_sheet(cases_ab, OUT_DIR / "expert_response_sheet.xlsx")
    build_mapping(cases_ab, OUT_DIR / "researcher_mapping_CONFIDENTIAL.xlsx")

    print("\nDONE.")
    print(f"  LaTeX:    {tex_out}")
    print(f"  To build: cd {OUT_DIR} && pdflatex expert_evaluation_cases.tex")
    print(f"  Mapping:  researcher_mapping_CONFIDENTIAL.xlsx  <-- keep private")


if __name__ == "__main__":
    main()
